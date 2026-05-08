#    -*- coding: UTF-8 -*-
#   @Author:   KingCorner
#   @Time:     2026/4/14 22:00
#   @FileRole: 安全预检与文件加载 (XML维数解析, 驱动路由)

import zipfile
import re
import datetime
import io
from typing import Union, BinaryIO, Tuple
from llm_excel_parser.utils.logger_module import get_logger
from llm_excel_parser.core.enums import ExcelFormat
from llm_excel_parser.config import default_config
from llm_excel_parser.core.exceptions import (
    ExcelParserBaseException,
    OverDimensionError,
    UnsupportedFormatError,
    InvalidInputTypeError,
    MissingDependencyError,
    FileCorruptedError
)

logger = get_logger("phase1_loader")
InputType = Union[str, bytes, BinaryIO]


class FileFormatSniffer:
    """内部工具类1：专职负责输入流归一化与文件魔数嗅探"""

    @staticmethod
    def detect_and_normalize(source: InputType) -> Tuple[ExcelFormat, BinaryIO]:
        file_obj = None
        if isinstance(source, str):
            with open(source, 'rb') as f:
                file_obj = io.BytesIO(f.read())
        elif isinstance(source, bytes):
            file_obj = io.BytesIO(source)
        elif hasattr(source, 'read'):
            file_obj = source
        else:
            raise InvalidInputTypeError("不受支持的输入类型，应为 str(路径), bytes 或 BytesIO")

        file_obj.seek(0)
        magic_number = file_obj.read(8)
        file_obj.seek(0)

        # ZIP (xlsx) -> PK.. => 50 4B 03 04
        if magic_number.startswith(b'\x50\x4B\x03\x04'):
            return ExcelFormat.XLSX, file_obj
        # OLE2 (xls) -> D0 CF 11 E0...
        elif magic_number.startswith(b'\xD0\xCF\x11\xE0'):
            return ExcelFormat.XLS, file_obj
        else:
            return ExcelFormat.UNKNOWN, file_obj


class DimensionChecker:
    """内部工具类2：专职负责维度安全检查 (防止 OOM 炸弹)"""

    @staticmethod
    def check_xlsx(file_obj: BinaryIO, max_rows: int, max_cols: int):
        try:
            with zipfile.ZipFile(file_obj, 'r') as zf:
                sheet_files = [f for f in zf.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
                for sheet_file in sheet_files:
                    with zf.open(sheet_file) as f:
                        content = f.read(1024)
                        match = re.search(rb'<dimension ref="([A-Z]+\d+):([A-Z]+\d+)"', content)
                        if not match:
                            continue

                        range_str = match.group(2).decode('utf-8')
                        col_str = "".join(filter(str.isalpha, range_str))
                        row_str = "".join(filter(str.isdigit, range_str))

                        col_num = sum(
                            (ord(c) - ord('A') + 1) * (26 ** i) for i, c in enumerate(reversed(col_str.upper())))
                        row_num = int(row_str) if row_str else 1

                        if row_num > max_rows or col_num > max_cols:
                            raise OverDimensionError(
                                f"安全触发: 工作表 {sheet_file} 行数({row_num})或列数({col_num})超出系统限制"
                            )
        except zipfile.BadZipFile:
            raise UnsupportedFormatError("XLSX 文件已损坏无法解压")
        finally:
            file_obj.seek(0)

    @staticmethod
    def check_xls(file_obj: BinaryIO, max_rows: int, max_cols: int):
        try:
            import xlrd
        except ImportError:
            return

        file_contents = file_obj.read()
        try:
            wb = xlrd.open_workbook(file_contents=file_contents, on_demand=True)
            for s_name in wb.sheet_names():
                sheet = wb.sheet_by_name(s_name)
                if sheet.nrows > max_rows or sheet.ncols > max_cols:
                    wb.release_resources()
                    raise OverDimensionError(f"安全触发: 工作表 {s_name} 的维度超出限制")
            wb.release_resources()
        except OverDimensionError as e:
            raise e
        except Exception as e:
            logger.warning(f"XLS维度校验异常，予以放行: {e}")
        finally:
            file_obj.seek(0)


class XlsToXlsxConverter:
    """内部工具类3：专职负责将老旧 xls 在内存中迁移到 openpyxl 结构"""

    @staticmethod
    def convert_in_memory(file_contents: bytes, include_hidden_rows: bool = False):
        logger.info(f"在内存中转换 XLS 为 XLSX 流水线对象, 源流大小: {len(file_contents)} bytes")
        try:
            import xlrd
            from openpyxl import Workbook
        except ImportError as e:
            raise MissingDependencyError(
                "处理 .xls 文件需要同时安装 'xlrd' 和 'openpyxl' (pip install xlrd openpyxl)") from e

        try:
            xls_workbook = xlrd.open_workbook(file_contents=file_contents, formatting_info=True)
        except Exception as e:
            raise FileCorruptedError(f"底层引擎(xlrd)无法读取 XLS 文件, 可能已损坏或加密: {e}") from e

        mem_wb = Workbook()
        if mem_wb.worksheets:
            mem_wb.remove(mem_wb.active)

        for xls_sheet in xls_workbook.sheets():
            ws = mem_wb.create_sheet(title=xls_sheet.name)

            if xls_sheet.visibility != 0:
                ws.sheet_state = 'hidden'

            hidden_rows_idx = set()
            for row_idx in range(xls_sheet.nrows):
                row_info = xls_sheet.rowinfo_map.get(row_idx)
                if row_info and getattr(row_info, 'hidden', False):
                    hidden_rows_idx.add(row_idx)

            if hidden_rows_idx:
                logger.debug(f"Sheet '{xls_sheet.name}' 检测到 {len(hidden_rows_idx)} 隐藏行。")

            xlsx_row_mapping = {}
            current_xlsx_row = 1

            for row in range(xls_sheet.nrows):
                if not include_hidden_rows and row in hidden_rows_idx:
                    continue

                xlsx_row_mapping[row] = current_xlsx_row
                if include_hidden_rows and row in hidden_rows_idx:
                    ws.row_dimensions[current_xlsx_row].hidden = True

                for col in range(xls_sheet.ncols):
                    cell_val = xls_sheet.cell_value(row, col)
                    cell_type = xls_sheet.cell_type(row, col)

                    if cell_type == xlrd.XL_CELL_DATE:
                        try:
                            dt_tuple = xlrd.xldate_as_tuple(cell_val, xls_workbook.datemode)
                            cell_val = datetime.datetime(*dt_tuple)
                        except Exception:
                            pass
                    elif cell_type == xlrd.XL_CELL_BOOLEAN:
                        cell_val = bool(cell_val)
                    elif cell_type in (xlrd.XL_CELL_ERROR, xlrd.XL_CELL_EMPTY):
                        cell_val = None

                    ws.cell(row=current_xlsx_row, column=col + 1, value=cell_val)

                current_xlsx_row += 1

            try:
                for min_row, max_row, min_col, max_col in xls_sheet.merged_cells:
                    if not include_hidden_rows:
                        valid_rows = [r for r in range(min_row, max_row) if r in xlsx_row_mapping]
                        if not valid_rows:
                            continue
                        start_r = xlsx_row_mapping[min(valid_rows)]
                        end_r = xlsx_row_mapping[max(valid_rows)]
                        start_c = min_col + 1
                        end_c = max_col
                        if start_r < end_r or start_c < end_c:
                            ws.merge_cells(start_row=start_r, start_column=start_c, end_row=end_r, end_column=end_c)
                    else:
                        ws.merge_cells(start_row=min_row + 1, start_column=min_col + 1, end_row=max_row,
                                       end_column=max_col)
            except AttributeError:
                pass
            except Exception as e:
                logger.debug(f"合并单元格解析异常: {e}")

        xls_workbook.release_resources()
        if not mem_wb.worksheets:
            mem_wb.create_sheet(title="Sheet1")

        logger.info(f"内存转换完成, 生成纯净 openpyxl 对象以供渲染")
        return mem_wb


class SecureLoader:
    """
    Phase 1 核心主类 (门面模式)
    负责串联：嗅探 -> 预检 -> 路由加载 整个流程
    """

    @classmethod
    def check_dimensions_and_route(cls, source: InputType,
                                   config_params: dict = None) -> list:
        config_params = config_params or {}
        max_rows = config_params.get("max_rows", default_config.MAX_ROW_LIMIT)
        max_cols = config_params.get("max_cols", default_config.MAX_COL_LIMIT)
        include_hidden_rows = config_params.get("include_hidden_rows", default_config.INCLUDE_HIDDEN_ROWS)
        include_hidden_sheets = config_params.get("include_hidden_sheets", default_config.INCLUDE_HIDDEN_SHEETS)

        fmt, file_obj = FileFormatSniffer.detect_and_normalize(source)
        if fmt == ExcelFormat.UNKNOWN:
            raise UnsupportedFormatError("未知的表格格式，仅支持xls和xlsx(以及二进制流)")
        if fmt == ExcelFormat.XLSX:
            DimensionChecker.check_xlsx(file_obj, max_rows, max_cols)
        else:
            DimensionChecker.check_xls(file_obj, max_rows, max_cols)
        file_obj.seek(0)

        # 路由加载引擎
        wb_obj = None
        try:
            if fmt == ExcelFormat.XLSX:
                from openpyxl import load_workbook
                logger.info("启动 XLSX 驱动...")
                wb_obj = load_workbook(file_obj, data_only=True)
            else:
                logger.info("启动 XLS 内存转换流水线...")
                wb_obj = XlsToXlsxConverter.convert_in_memory(file_obj.read(), include_hidden_rows)
        except ExcelParserBaseException:
            raise
        except Exception as e:
            raise FileCorruptedError(f"文件加载失败，无法解析底层工作簿: {str(e)}") from e

        from llm_excel_parser.adapters.openpyxl_adapter import OpenpyxlWorksheetAdapter
        sheets = []
        for sheet in wb_obj.worksheets:
            # Phase 1.3: 根据配置决定是否跳过隐藏工作表
            if not include_hidden_sheets and sheet.sheet_state != 'visible':
                logger.debug(f"跳过隐藏工作表: '{sheet.title}' (state={sheet.sheet_state})")
                continue
            sheets.append(OpenpyxlWorksheetAdapter(sheet))

        logger.info(f"Phase 1 完成，共加载 {len(sheets)} 张工作表"
                    + ("（含隐藏）" if include_hidden_sheets else "（隐藏已过滤）"))
        return sheets
