#    -*- coding: UTF-8 -*-
#   @Author:   KingCorner
#   @Time:     2026/4/14 22:01
#   @FileRole: 数据渲染层 - 将 Excel 数据块转换为脱离引擎的二维字典 / 数组

from typing import List, Dict, Any, Tuple
from openpyxl.utils import get_column_letter  #Todo：当前底层使用的是 openpyxl, 未来如果需要支持其它引擎, 则需要解耦
from llm_excel_parser.utils.logger_module import get_logger
from llm_excel_parser.core.enums import MergeAction
from llm_excel_parser.core.datatypes import BoundingBox
from llm_excel_parser.core.exceptions import DataRenderError
from llm_excel_parser.utils.formatters import format_cell_value  # 假设已实现的值格式化器

logger = get_logger("phase3_renderer")


class DataRenderer:
    @staticmethod
    def render_box(ws, box: BoundingBox, action: MergeAction, ignore_hidden: bool = True) -> List[List[Any]]:
        """
        阶段 3: 提取 Box 区域的真实数据并处理合并单元格

        :param ws: worksheet 对象 (openpyxl)
        :param box: Phase 2 产出的表格包围盒
        :param action: 处理合并单元格的策略枚举
        :return: 二维 List，代表该Box内所有可见单元格经过策略渲染后的最终值
        """
        if box.max_row == 0 or box.max_col == 0:
            return []
        try:
            # 1. 预计算可见行/列
            visible_rows = []
            for r in range(box.min_row, box.max_row + 1):
                # 如果开启了忽略隐藏行，且当前行是隐藏的，才跳过
                if ignore_hidden and (r in ws.row_dimensions and ws.row_dimensions[r].hidden):
                    continue
                visible_rows.append(r)

            visible_cols = []
            for c in range(box.min_col, box.max_col + 1):
                col_letter = get_column_letter(c)
                # 同理，处理隐藏列
                if ignore_hidden and (col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].hidden):
                    continue
                visible_cols.append(c)
            if not visible_rows or not visible_cols:
                return []

            # 2. 从全局合并单元格中，过滤出仅在这个 Box 内产生交集的拓扑关系
            # （性能优化点：避免 O(M) 级别的无意义循环）
            merged_dict: Dict[Tuple[int, int], Dict[str, Any]] = {}
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds

                # 矩阵碰撞检测: 如果该合并单元格和当前 Box 没有任何交集，直接跳过
                if max_col < box.min_col or min_col > box.max_col or \
                        max_row < box.min_row or min_row > box.max_row:
                    continue
                # 提取该合并区域的主格原值 (左上角)
                master_val = format_cell_value(ws.cell(row=min_row, column=min_col).value)
                # 将属于交集范围的合并单元格记录在字典内，供渲染时快速 O(1) 索引
                for r in range(max(min_row, box.min_row), min(max_row, box.max_row) + 1):
                    for c in range(max(min_col, box.min_col), min(max_col, box.max_col) + 1):
                        is_top_left = (r == min_row and c == min_col)
                        merged_dict[(r, c)] = {
                            "master_val": master_val,
                            "is_top_left": is_top_left
                        }
            # 3. 严格构建最终二维数据组 List[List[Any]]
            rendered_matrix: List[List[Any]] = []

            for r in visible_rows:
                row_data = []
                for c in visible_cols:
                    raw_val = ws.cell(row=r, column=c).value
                    formatted_val = format_cell_value(raw_val)
                    final_val = formatted_val
                    # 4. 执行特定的 MergeAction 策略 (Merge Cell Resolution)
                    if (r, c) in merged_dict:
                        merge_info = merged_dict[(r, c)]
                        if action == MergeAction.FILL_FORWARD:
                            final_val = merge_info["master_val"]
                        elif action == MergeAction.TOP_LEFT:
                            # 仅保留在整个合并区域最左上角的那一点
                            final_val = merge_info["master_val"] if merge_info["is_top_left"] else None
                        elif action == MergeAction.TAG:
                            master = merge_info["master_val"]
                            final_val = f"[MERGED_REF({master})]" if master is not None else "[MERGED_REF(None)]"
                    row_data.append(final_val)
                rendered_matrix.append(row_data)
            logger.debug(f"Box({box.range_str}) 数据渲染完成，矩阵维度: "
                         f"行数={len(rendered_matrix)}, 宽={len(rendered_matrix[0]) if rendered_matrix else 0}")
            return rendered_matrix
        except Exception as e:
            # 防止第三方底层库错误污染上层流水线，包成系统能识别的统一异常
            logger.error(f"渲染 Box({box.range_str}) 失败! Error: {str(e)}")
            raise DataRenderError(f"Phase 3 数据提取异常: 提取 box {box.range_str} 失败. 具体原因: {str(e)}") from e
