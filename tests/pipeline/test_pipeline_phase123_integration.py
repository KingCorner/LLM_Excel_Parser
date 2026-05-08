#    -*- coding: UTF-8 -*-
#   @Author:   KingCorner
#   @Time:     2026/4/20 17:38
#   @FileRole: 对phase1~3的单元测试


import os
import tempfile
import pytest
import openpyxl

from llm_excel_parser.pipeline.phase1_loader import SecureLoader
from llm_excel_parser.pipeline.phase2_detector import StructureDetector
from llm_excel_parser.pipeline.phase3_renderer import DataRenderer
from llm_excel_parser.core.enums import MergeAction

from llm_excel_parser.adapters.openpyxl_adapter import OpenpyxlWorksheetAdapter
from llm_excel_parser.core.interfaces import BaseWorksheet


@pytest.fixture
def complex_excel_file(tmp_path):
    """
    动态生成一个真实的 Excel 物理文件，用于集成测试。
    该表格包含两个独立的数据块（被空行隔开），并含有合并单元格。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IntegrationSheet"

    # 表格块 1: 位于左上角 (A1:C3)
    # 包含一个合并单元格 A1:B1 (跨两列)
    ws['A1'] = "合并标头"
    ws.merge_cells('A1:B1')
    ws['C1'] = "独立标头"

    ws['A2'] = "数据A2"
    ws['B2'] = "数据B2"
    ws['C2'] = "数据C2"

    ws['A3'] = "数据A3"
    ws['B3'] = "数据B3"
    ws['C3'] = "数据C3"

    # 空出第4、第5行，形成物理断层，用于触发 Phase 2 的切表逻辑

    # 表格块 2: 位于偏右下方 (B6:C7)

    ws['B6'] = "表2标题1"
    ws['C6'] = "表2标题2"
    ws['B7'] = "内容1"
    ws['C7'] = "内容2"

    # 利用 pytest 的 tmp_path 构建路径
    filepath = tmp_path / "integration_test.xlsx"
    wb.save(filepath)
    return str(filepath)


def test_main_pipeline_happy_path(complex_excel_file):
    """
    主干链路集成测试：
    验证从加载文件 -> 识别多表格边界 -> 渲染二维数据的完整链路
    """

    # Phase 1: 安全预检与文件加载
    # 送入真实的 xlsx 物理路径
    worksheets = SecureLoader.check_dimensions_and_route(complex_excel_file)

    assert len(worksheets) == 1, "应该成功解析出一个 Worksheet 对象"
    ws = worksheets[0]
    assert ws.title == "IntegrationSheet", "Phase 1: 工作表读取或适配失败"

    assert isinstance(ws, OpenpyxlWorksheetAdapter)

    # Phase 2: 结构探测层 (切表)
    boxes = StructureDetector.detect_tables(ws, max_empty_rows=1)

    # 验证是否成功跨越空行，识别出两个独立的表格包围盒
    assert len(boxes) == 2, "Phase 2: 未能正确识别出 2 个独立的表格块"

    # 按行号排序以便后续断言 (由于底层探测算法可能不保证先后顺序)
    boxes = sorted(boxes, key=lambda b: b.min_row)
    box1 = boxes[0]  # 预期是 A1:C3
    box2 = boxes[1]  # 预期是 B6:C7

    assert box1.min_row == 1 and box1.max_row == 3 and box1.min_col == 1 and box1.max_col == 3, "表 1 坐标推断错误"
    assert box2.min_row == 6 and box2.max_row == 7 and box2.min_col == 2 and box2.max_col == 3, "表 2 坐标推断错误"

    # Phase 3: 数据渲染层 (二维数组提取 & 合并单元格消解)

    # 场景 1：使用 FILL_FORWARD 策略渲染表 1 (验证 A1:B1 的合并单元格是否被平铺补齐)
    rendered_table_1_fill = DataRenderer.render_box(ws, box1, action=MergeAction.FILL_FORWARD)

    expected_table_1_fill = [
        # 注意："合并标头" 应该在第 1 行 第 2 列 (B1) 被自动复制填充了
        ["合并标头", "合并标头", "独立标头"],
        ["数据A2", "数据B2", "数据C2"],
        ["数据A3", "数据B3", "数据C3"]
    ]

    assert rendered_table_1_fill == expected_table_1_fill, "Phase 3: 向下向右填充合并单元格策略失效"

    # 场景 2：使用 TOP_LEFT 策略渲染表 1 (验证非左上角是否留空，这里假设你的策略返回 None 或 "")
    rendered_table_1_topleft = DataRenderer.render_box(ws, box1, action=MergeAction.TOP_LEFT)

    # 取第一行，验证第二个元素 (B1) 不是"合并标头"
    top_left_header_row = rendered_table_1_topleft[0]
    assert top_left_header_row[0] == "合并标头"
    assert top_left_header_row[1] is None or top_left_header_row[1] == "", "Phase 3: TOP_LEFT 时非主格未妥善置空"
    assert top_left_header_row[2] == "独立标头"

    # 场景 3：渲染表 2，确保坐标偏移情况下数据截取完全正确 (只取 B6:C7 的内容)
    rendered_table_2 = DataRenderer.render_box(ws, box2, action=MergeAction.FILL_FORWARD)

    expected_table_2 = [
        ["表2标题1", "表2标题2"],
        ["内容1", "内容2"]
    ]
    assert rendered_table_2 == expected_table_2, "Phase 3: 偏移表格内容提取错误"

    print("\n[✔] 流水线 Phase 1 -> Phase 2 -> Phase 3 贯通集成测试全部通过！")


# ===== Phase 1.3: 隐藏工作表过滤测试 =====

@pytest.fixture
def workbook_with_hidden_sheet(tmp_path):
    """生成一个包含 1 张可见 Sheet 和 1 张隐藏 Sheet 的 xlsx 文件"""
    wb = openpyxl.Workbook()

    visible_ws = wb.active
    visible_ws.title = "Visible"
    visible_ws['A1'] = "可见内容"

    hidden_ws = wb.create_sheet(title="Hidden")
    hidden_ws.sheet_state = 'hidden'
    hidden_ws['A1'] = "隐藏内容"

    filepath = tmp_path / "hidden_sheet_test.xlsx"
    wb.save(filepath)
    return str(filepath)


def test_hidden_sheet_filtered_by_default(workbook_with_hidden_sheet):
    """默认情况下隐藏工作表应被过滤，只返回可见 Sheet"""
    worksheets = SecureLoader.check_dimensions_and_route(workbook_with_hidden_sheet)

    assert len(worksheets) == 1
    assert worksheets[0].title == "Visible"


def test_hidden_sheet_included_when_flag_true(workbook_with_hidden_sheet):
    """include_hidden_sheets=True 时隐藏工作表应被纳入"""
    worksheets = SecureLoader.check_dimensions_and_route(
        workbook_with_hidden_sheet,
        config_params={"include_hidden_sheets": True}
    )

    assert len(worksheets) == 2
    titles = {ws.title for ws in worksheets}
    assert titles == {"Visible", "Hidden"}


# ===== Phase 2: 隐藏行/列对结构探测的影响 =====

@pytest.fixture
def workbook_with_hidden_row_col(tmp_path):
    """
    生成一个包含隐藏行/列的 xlsx。
    布局：
      行1: A1=标题A  B1=标题B  C1=标题C
      行2: A2=数据1  B2=数据2  C2=数据3  ← 整行隐藏
      行3: A3=数据4  B3=数据5  C3=数据6
    列B整列隐藏。

    ignore_hidden=True 时，Phase 2 布尔矩阵只应看到：
      可见行 1, 3；可见列 A(1), C(3)
    → 仍是一个连通块，但列坐标只有 1 和 3
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HiddenRowCol"

    for r in range(1, 4):
        ws.cell(row=r, column=1, value=f"A{r}")
        ws.cell(row=r, column=2, value=f"B{r}")
        ws.cell(row=r, column=3, value=f"C{r}")

    ws.row_dimensions[2].hidden = True
    ws.column_dimensions['B'].hidden = True

    filepath = tmp_path / "hidden_row_col_test.xlsx"
    wb.save(filepath)
    return str(filepath)


def test_phase2_excludes_hidden_rows_and_cols(workbook_with_hidden_row_col):
    """ignore_hidden=True 时，Phase 2 的布尔矩阵不应纳入隐藏行/列的单元格"""
    worksheets = SecureLoader.check_dimensions_and_route(workbook_with_hidden_row_col)
    ws = worksheets[0]

    boxes = StructureDetector.detect_tables(ws, ignore_hidden=True)

    assert len(boxes) == 1, "可见单元格应归并为一个连通块"
    box = boxes[0]

    # 隐藏行 2 应不在范围内（行范围应为 1~3，但行2本身被跳过）
    # 连通域包围盒涵盖 row 1 和 row 3，所以 min_row=1, max_row=3
    assert box.min_row == 1
    assert box.max_row == 3
    # 隐藏列 B(2) 应不在范围内，只有列 A(1) 和 C(3)
    assert box.min_col == 1
    assert box.max_col == 3


def test_phase2_includes_hidden_rows_and_cols_when_flag_false(workbook_with_hidden_row_col):
    """ignore_hidden=False 时，隐藏行/列的单元格应被纳入布尔矩阵"""
    worksheets = SecureLoader.check_dimensions_and_route(workbook_with_hidden_row_col)
    ws = worksheets[0]

    boxes = StructureDetector.detect_tables(ws, ignore_hidden=False)

    assert len(boxes) == 1
    box = boxes[0]
    # 所有行列均纳入，包围盒应覆盖 1~3 行、1~3 列
    assert box.min_row == 1 and box.max_row == 3
    assert box.min_col == 1 and box.max_col == 3


# ===== Phase 3: 隐藏主格合并区域的值隔离测试 =====

@pytest.fixture
def workbook_hidden_row_master_merge(tmp_path):
    """
    行1隐藏（主格 A1="HIDDEN_MASTER"，与 A2 合并），行2可见（从格），行3可见独立格 A3="VISIBLE"
    用于验证隐藏行主格的值不应通过策略渗入可见从格 A2
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HiddenRowMasterTest"
    ws['A1'] = "HIDDEN_MASTER"
    ws.merge_cells('A1:A2')
    ws['A3'] = "VISIBLE"
    ws.row_dimensions[1].hidden = True
    filepath = tmp_path / "hidden_row_master_merge.xlsx"
    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def workbook_hidden_col_master_merge(tmp_path):
    """
    列A隐藏（主格 A1="HIDDEN_MASTER"，与 B1 合并），列B可见（从格），行2两列均可见
    用于验证隐藏列主格的值不应通过策略渗入可见从格 B1
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HiddenColMasterTest"
    ws['A1'] = "HIDDEN_MASTER"
    ws.merge_cells('A1:B1')
    ws['A2'] = "VA2"
    ws['B2'] = "VB2"
    ws.column_dimensions['A'].hidden = True
    filepath = tmp_path / "hidden_col_master_merge.xlsx"
    wb.save(filepath)
    return str(filepath)


def test_phase3_hidden_row_master_no_value_leak(workbook_hidden_row_master_merge):
    """ignore_hidden=True 时，隐藏行主格的值不应通过 FILL_FORWARD 渗入可见从格"""
    worksheets = SecureLoader.check_dimensions_and_route(workbook_hidden_row_master_merge)
    ws = worksheets[0]

    boxes = StructureDetector.detect_tables(ws, ignore_hidden=True)
    assert len(boxes) == 1
    box = boxes[0]

    rendered = DataRenderer.render_box(ws, box, action=MergeAction.FILL_FORWARD, ignore_hidden=True)

    # 可见行: row2(从格), row3; 只有 A 列
    assert len(rendered) == 2
    # row2 从格：主格在隐藏行，应为空而非 "HIDDEN_MASTER"
    assert rendered[0][0] != "HIDDEN_MASTER", "隐藏行主格的值不应渗入可见从格"
    assert rendered[0][0] is None or rendered[0][0] == ""
    # row3 独立格
    assert rendered[1][0] == "VISIBLE"


def test_phase3_hidden_col_master_no_value_leak(workbook_hidden_col_master_merge):
    """ignore_hidden=True 时，隐藏列主格的值不应通过 FILL_FORWARD 渗入可见从格"""
    worksheets = SecureLoader.check_dimensions_and_route(workbook_hidden_col_master_merge)
    ws = worksheets[0]

    boxes = StructureDetector.detect_tables(ws, ignore_hidden=True)
    assert len(boxes) == 1
    box = boxes[0]

    rendered = DataRenderer.render_box(ws, box, action=MergeAction.FILL_FORWARD, ignore_hidden=True)

    # 可见列: col B(2); 可见行: row1, row2
    assert len(rendered) == 2
    # row1 B1 从格：主格在隐藏列 A，应为空
    assert rendered[0][0] != "HIDDEN_MASTER", "隐藏列主格的值不应渗入可见从格"
    assert rendered[0][0] is None or rendered[0][0] == ""
    # row2 B2 独立格
    assert rendered[1][0] == "VB2"


def test_phase3_hidden_master_included_when_not_ignoring(workbook_hidden_row_master_merge):
    """ignore_hidden=False 时，隐藏主格的值正常通过策略传播到从格"""
    worksheets = SecureLoader.check_dimensions_and_route(workbook_hidden_row_master_merge)
    ws = worksheets[0]

    boxes = StructureDetector.detect_tables(ws, ignore_hidden=False)
    assert len(boxes) == 1
    box = boxes[0]

    rendered = DataRenderer.render_box(ws, box, action=MergeAction.FILL_FORWARD, ignore_hidden=False)

    # ignore_hidden=False: 行1(主格)+行2(从格)+行3；FILL_FORWARD 应将 "HIDDEN_MASTER" 传播到 A2
    assert len(rendered) == 3
    assert rendered[0][0] == "HIDDEN_MASTER"   # A1 主格
    assert rendered[1][0] == "HIDDEN_MASTER"   # A2 从格（FILL_FORWARD）
    assert rendered[2][0] == "VISIBLE"         # A3 独立格
