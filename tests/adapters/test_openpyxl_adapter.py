#    -*- coding: UTF-8 -*-
#   @Author:   KingCorner
#   @Time:     2026/4/20 17:30
#   @FileRole: 对openpyxl_adapter的单元测试


import pytest
import openpyxl
from openpyxl.styles import Font, PatternFill
from unittest.mock import Mock, patch, PropertyMock

from llm_excel_parser.adapters.openpyxl_adapter import OpenpyxlWorksheetAdapter


@pytest.fixture
def real_workbook():
    """提供一个真实的 openpyxl Workbook 对象用于构建测试场景"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # 填充 3x3 的基础数据
    # A1(1,1)="1-1", B1(1,2)="1-2" ... C3(3,3)="3-3"
    for r in range(1, 4):
        for c in range(1, 4):
            ws.cell(row=r, column=c, value=f"{r}-{c}")

    # 设置合并单元格: A1:B2 (row1~row2, col1~col2)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)

    # 隐藏第2行
    ws.row_dimensions[2].hidden = True

    # 隐藏第B列 (即第2列)
    ws.column_dimensions['B'].hidden = True

    return wb, ws


@pytest.fixture
def adapter(real_workbook):
    """返回初始化好的 Adapter 实例"""
    _, ws = real_workbook
    return OpenpyxlWorksheetAdapter(ws)


class TestOpenpyxlWorksheetAdapter:
    """OpenpyxlWorksheetAdapter 的单元测试套件"""

    def test_property_title(self, adapter):
        """测试 sheet 名称获取"""
        assert adapter.title == "TestSheet"

    def test_property_max_dimensions(self, adapter):
        """测试最大行列数获取"""
        # 我们在 fixture 中填充了 3x3 的数据
        assert adapter.max_dimensions == (3, 3)

    def test_get_cell_value(self, adapter):
        """测试单元格值获取"""
        assert adapter.get_cell_value(1, 1) == "1-1"
        assert adapter.get_cell_value(3, 2) == "3-2"
        # 测试空单元格 (超出填充范围)
        assert adapter.get_cell_value(4, 4) is None

    def test_iter_rows(self, adapter):
        """测试行迭代器功能"""
        # 获取 Row 1 到 Row 2, Col 1 到 Col 2 的数据 (即 A1:B2)
        generator = adapter.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2)
        rows = list(generator)

        assert len(rows) == 2

        # A1:B2已被合并，故除了 A1('1-1') 之外，B1、A2、B2 的 raw value 均为 None
        assert rows[0] == ("1-1", None)
        assert rows[1] == (None, None)

    def test_get_merged_regions(self, adapter):
        """测试合并单元格区域获取"""
        regions = adapter.get_merged_regions()

        # 期望返回格式: (min_row, min_col, max_row, max_col)
        # 我们在 fixture 中合并了 A1:B2 -> (1, 1, 2, 2)
        assert len(regions) == 1
        assert regions[0] == (1, 1, 2, 2)

    def test_get_merged_regions_no_attr(self):
        """测试没有 merged_cells 属性时的容错处理"""
        mock_ws = Mock()
        del mock_ws.merged_cells  # 强制移除属性
        mock_adapter = OpenpyxlWorksheetAdapter(mock_ws)

        assert mock_adapter.get_merged_regions() == []

    def test_is_row_hidden_true(self, adapter):
        """测试隐藏行的判断 (真)"""
        assert adapter.is_row_hidden(2) is True

    def test_is_row_hidden_false(self, adapter):
        """测试隐藏行的判断 (假)"""
        assert adapter.is_row_hidden(1) is False
        assert adapter.is_row_hidden(999) is False  # 不存在的行默认不隐藏

    def test_is_row_hidden_no_attr(self):
        """测试没有 row_dimensions 属性时的容错处理"""
        mock_ws = Mock()
        del mock_ws.row_dimensions
        mock_adapter = OpenpyxlWorksheetAdapter(mock_ws)

        assert mock_adapter.is_row_hidden(1) is False

    # 我们需要 patch Adapter 内部导入的 col_idx_to_letter，使其返回真实的字母
    # 因为在没有加载整个项目环境的情况下，保证测试独立性是很重要的
    @patch('llm_excel_parser.adapters.openpyxl_adapter.col_idx_to_letter')
    def test_is_col_hidden_true(self, mock_col_idx_to_letter, adapter):
        """测试隐藏列的判断 (真)"""
        mock_col_idx_to_letter.return_value = 'B'
        assert adapter.is_col_hidden(2) is True
        mock_col_idx_to_letter.assert_called_with(2)

    @patch('llm_excel_parser.adapters.openpyxl_adapter.col_idx_to_letter')
    def test_is_col_hidden_false(self, mock_col_idx_to_letter, adapter):
        """测试隐藏列的判断 (假)"""
        mock_col_idx_to_letter.return_value = 'A'
        assert adapter.is_col_hidden(1) is False

        mock_col_idx_to_letter.return_value = 'Z'
        assert adapter.is_col_hidden(26) is False

    @patch('llm_excel_parser.adapters.openpyxl_adapter.col_idx_to_letter')
    def test_is_col_hidden_no_attr(self, mock_col_idx_to_letter):
        """测试没有 column_dimensions 属性时的容错处理"""
        mock_ws = Mock()
        del mock_ws.column_dimensions
        mock_adapter = OpenpyxlWorksheetAdapter(mock_ws)

        assert mock_adapter.is_col_hidden(1) is False
        # 如果根本没有这个属性，就不应该走到转换列名字母那一步
        mock_col_idx_to_letter.assert_not_called()


# ===== get_cell_style 测试 =====
@pytest.fixture
def styled_workbook():
    """构建带有各种样式组合的真实 Workbook，用于 get_cell_style 测试。
    单元格布局：                                                                                                                                                   
      A1 - 无样式 (空白)                                                                                                                                           
      A2 - 粗体字体                                                                                                                                                
      A3 - 纯色背景填充 (红色 FF0000)                                                                                                                              
      A4 - 粗体 + 背景                                                                                                                                             
      A5 - 填充类型为 None (透明，不计入背景)                                                                                                                      
      A6 - 填充 fgColor.rgb == '00000000' (openpyxl 透明色，不计入背景)                                                                                            
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StyleSheet"

    ws['A1'].value = "plain"

    ws['A2'].value = "bold"
    ws['A2'].font = Font(bold=True)

    ws['A3'].value = "bg"
    ws['A3'].fill = PatternFill(fill_type='solid', fgColor='FF0000')

    ws['A4'].value = "bold+bg"
    ws['A4'].font = Font(bold=True)
    ws['A4'].fill = PatternFill(fill_type='solid', fgColor='00FF00')

    ws['A5'].value = "fill_none"
    ws['A5'].fill = PatternFill(fill_type=None)

    # fgColor='00000000' 是 openpyxl 默认的"无色"RGB，应被视为透明
    ws['A6'].value = "transparent_rgb"
    ws['A6'].fill = PatternFill(fill_type='solid', fgColor='00000000')

    return wb, ws


@pytest.fixture
def style_adapter(styled_workbook):
    _, ws = styled_workbook
    return OpenpyxlWorksheetAdapter(ws)


class TestGetCellStyle:
    """get_cell_style 方法的单元测试套件"""

    def test_plain_cell_returns_empty_dict(self, style_adapter):
        """无样式的普通单元格应返回空字典"""
        assert style_adapter.get_cell_style(1, 1) == {}

    def test_bold_font_sets_is_bold(self, style_adapter):
        """粗体单元格应设置 is_bold=True"""
        style = style_adapter.get_cell_style(2, 1)
        assert style.get('is_bold') is True
        assert 'has_bg' not in style

    def test_non_bold_font_omits_is_bold(self, style_adapter):
        """非粗体单元格不应包含 is_bold 键"""
        style = style_adapter.get_cell_style(3, 1)
        assert 'is_bold' not in style

    def test_solid_fill_sets_has_bg(self, style_adapter):
        """纯色背景填充应设置 has_bg=True"""
        style = style_adapter.get_cell_style(3, 1)
        assert style.get('has_bg') is True

    def test_bold_and_bg_both_set(self, style_adapter):
        """同时有粗体和背景时，两个键均应出现"""
        style = style_adapter.get_cell_style(4, 1)
        assert style.get('is_bold') is True
        assert style.get('has_bg') is True

    def test_fill_type_none_omits_has_bg(self, style_adapter):
        """fill_type=None 的单元格不应被视为有背景"""
        style = style_adapter.get_cell_style(5, 1)
        assert 'has_bg' not in style

    def test_transparent_rgb_omits_has_bg(self, style_adapter):
        """fgColor.rgb == '00000000' 表示无填充色，不应设置 has_bg"""
        style = style_adapter.get_cell_style(6, 1)
        assert 'has_bg' not in style

    def test_font_exception_swallowed_fill_still_read(self):
        """font 属性访问抛出异常时，应静默吞掉，fill 仍正常读取"""
        mock_ws = Mock()
        mock_cell = Mock()
        mock_ws.cell.return_value = mock_cell

        # font 访问抛出异常
        type(mock_cell).font = PropertyMock(side_effect=Exception("font error"))

        # fill 正常，有背景
        mock_fill = Mock()
        mock_fill.fill_type = 'solid'
        mock_fill.fgColor.rgb = 'FF0000'
        type(mock_cell).fill = PropertyMock(return_value=mock_fill)

        adapter = OpenpyxlWorksheetAdapter(mock_ws)
        style = adapter.get_cell_style(1, 1)

        assert 'is_bold' not in style
        assert style.get('has_bg') is True

    def test_fill_exception_swallowed_font_still_read(self):
        """fill 属性访问抛出异常时，应静默吞掉，font 仍正常读取"""
        mock_ws = Mock()
        mock_cell = Mock()
        mock_ws.cell.return_value = mock_cell

        # font 正常，是粗体
        mock_font = Mock()
        mock_font.bold = True
        type(mock_cell).font = PropertyMock(return_value=mock_font)

        # fill 访问抛出异常
        type(mock_cell).fill = PropertyMock(side_effect=Exception("fill error"))

        adapter = OpenpyxlWorksheetAdapter(mock_ws)
        style = adapter.get_cell_style(1, 1)

        assert style.get('is_bold') is True
        assert 'has_bg' not in style

    def test_both_exceptions_returns_empty_dict(self):
        """font 和 fill 同时抛出异常时，应返回空字典，不崩溃"""
        mock_ws = Mock()
        mock_cell = Mock()
        mock_ws.cell.return_value = mock_cell

        type(mock_cell).font = PropertyMock(side_effect=Exception("font error"))
        type(mock_cell).fill = PropertyMock(side_effect=Exception("fill error"))

        adapter = OpenpyxlWorksheetAdapter(mock_ws)
        style = adapter.get_cell_style(1, 1)

        assert style == {}
